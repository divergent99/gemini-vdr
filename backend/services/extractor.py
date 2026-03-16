import os
import base64
import tempfile
from backend.logger import log

SUPPORTED = (".pdf", ".docx", ".xlsx", ".xls")
MAX_CHARS  = 40_000
PER_FILE   = 10_000


def _parse_file(path: str, label: str) -> str:
    ext = path.split(".")[-1].lower()
    try:
        if ext == "pdf":
            import fitz
            text = "\n".join(p.get_text() for p in fitz.open(path))
        elif ext == "docx":
            from docx import Document
            text = "\n".join(p.text for p in Document(path).paragraphs if p.text.strip())
        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            lines = []
            for sh in wb.sheetnames:
                ws = wb[sh]
                lines.append(f"[Sheet: {sh}]")
                for row in ws.iter_rows(values_only=True):
                    v = [str(x) for x in row if x is not None]
                    if v:
                        lines.append(" | ".join(v))
            text = "\n".join(lines)
        else:
            log(f"Skipping unsupported file: {label}", "WARN")
            return ""
        log(f"Extracted {len(text):,} chars from {label}", "OK")
        return f"\n\n{'='*60}\nDOC: {label}\n{'='*60}\n{text[:PER_FILE]}"
    except Exception as e:
        log(f"Failed to parse {label}: {e}", "ERR")
        return f"\n[Failed to parse {label}: {e}]"


def extract_from_folder(folder_path: str) -> tuple[str, list[str]]:
    """Extract text from all supported files in a folder."""
    folder_path = folder_path.strip().strip('"').strip("'")
    if not os.path.isdir(folder_path):
        raise ValueError(f"Folder not found: {folder_path}")
    files = [f for f in os.listdir(folder_path) if f.lower().endswith(SUPPORTED)]
    if not files:
        raise ValueError(f"No supported files (PDF/DOCX/XLSX) in: {folder_path}")
    log(f"Found {len(files)} document(s) in {folder_path}")
    chunks = [_parse_file(os.path.join(folder_path, f), f) for f in sorted(files)]
    combined = "\n".join(chunks)[:MAX_CHARS]
    log(f"Total extracted: {len(combined):,} chars", "OK")
    return combined, files


def extract_from_uploads(contents_list: list, filenames_list: list) -> str:
    """Extract text from base64-encoded uploaded files."""
    log(f"Extracting {len(filenames_list)} uploaded file(s)...")
    chunks = []
    for content, filename in zip(contents_list, filenames_list):
        ext = filename.split(".")[-1].lower()
        _, data = content.split(",", 1)
        file_bytes = base64.b64decode(data)
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        chunks.append(_parse_file(tmp_path, filename))
        os.unlink(tmp_path)
    combined = "\n".join(chunks)[:MAX_CHARS]
    log(f"Total extracted: {len(combined):,} chars", "OK")
    return combined


def scan_folder(folder_path: str) -> list[str]:
    """Return list of supported filenames in folder (no text extraction)."""
    folder_path = folder_path.strip().strip('"').strip("'")
    if not os.path.isdir(folder_path):
        return []
    return sorted(f for f in os.listdir(folder_path) if f.lower().endswith(SUPPORTED))
